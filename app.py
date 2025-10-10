import os, io, csv, base64,re
from datetime import datetime, date, timedelta

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, Response, jsonify, jsonify, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from sqlalchemy import text, func, case

try:
    import qrcode
except Exception:
    qrcode = None

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY") or "dev-key-change-me"
#app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///invoices.db"
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL") or "sqlite:///invoices.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config['JSON_AS_ASCII'] = False
db = SQLAlchemy(app)


# ---- Keep user filters (ym/start/end) across actions ----
from flask import make_response

def _current_filter_args():
    keys = ("ym", "start", "end", "status", "q")
    out = {}
    for k in keys:
        v = request.values.get(k) or request.args.get(k) or request.cookies.get(k)
        if v:
            out[k] = v
    return out

@app.after_request
def _remember_filters(resp):
    try:
        for k in ("ym", "start", "end", "status", "q"):
            v = request.values.get(k)
            if v:
                resp.set_cookie(k, v, max_age=7*24*3600, samesite="Lax")
    except Exception:
        pass
    return resp


# ---------------- Models ----------------
class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(16), nullable=False, default="employee")
    branch_number = db.Column(db.String(64), nullable=True)
    min_visible_date = db.Column(db.Date, nullable=True)
    is_admin = db.Column(db.Boolean, nullable=False, default=False)

    def set_password(self, pw: str): self.password_hash = generate_password_hash(pw)
    def check_password(self, pw: str) -> bool: return check_password_hash(self.password_hash, pw)

class Pricing(db.Model):
    __tablename__ = "pricing"
    id = db.Column(db.Integer, primary_key=True)
    unit_price = db.Column(db.Float, nullable=False, default=0.0)
    fee_20 = db.Column(db.Float, nullable=False, default=0.0)
    fee_15 = db.Column(db.Float, nullable=False, default=0.0)   # 15A
    fee_10 = db.Column(db.Float, nullable=False, default=0.0)
    fee_5  = db.Column(db.Float, nullable=False, default=0.0)
    currency_code = db.Column(db.String(8), nullable=False, default='LBP')
    usd_rate = db.Column(db.Float, nullable=False, default=89700.0)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    def fee_for_amp(self, amps: int) -> float:
        return float(
            self.fee_20 if amps == 20 else
            self.fee_15 if amps == 15 else
            self.fee_10 if amps == 10 else
            self.fee_5  if amps == 5  else
            0.0
        )

class Invoice(db.Model):
    __tablename__ = "invoices"
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(64), unique=True, nullable=False, index=True)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    customer_name = db.Column(db.String(255), nullable=False)
    meter_number = db.Column(db.String(64), nullable=True)
    branch_number = db.Column(db.String(64), nullable=False, index=True)
    subscription_amps = db.Column(db.Integer, nullable=True)
    prev_reading = db.Column(db.Integer, nullable=False, default=0)
    curr_reading = db.Column(db.Integer, nullable=False, default=0)
    unit_price = db.Column(db.Float, nullable=False, default=0.0)
    subscription_fee = db.Column(db.Float, nullable=False, default=0.0)
    is_paid = db.Column(db.Boolean, nullable=False, default=False)
    kwh_used = db.Column(db.Integer, nullable=False, default=0)
    energy_cost = db.Column(db.Float, nullable=False, default=0.0)
    month_cost = db.Column(db.Float, nullable=False, default=0.0)
    total_due = db.Column(db.Float, nullable=False, default=0.0)

# ---------------- Helpers ----------------

@app.template_filter("money")
def money(val):
    try:
        p = get_pricing()
        rate = float(p.usd_rate or 1.0)
        if p.currency_code == 'USD':
            amt = (float(val or 0.0)) / (rate if rate else 1.0)
            return f"${amt:,.2f}"
        else:
            amt = float(val or 0.0)
            return f"{int(round(amt)):,} ل.ل"
    except Exception:
        try:
            return f"{float(val):,.2f}"
        except Exception:
            return str(val)


@app.template_filter("money_lbp")
def money_lbp(val):
    try:
        amt = float(val or 0.0)
        return f"{int(round(amt)):,} ل.ل"
    except Exception:
        return str(val)

@app.template_filter("money_usd")
def money_usd(val):
    try:
        p = get_pricing()
        rate = float(p.usd_rate or 1.0) or 1.0
        amt = (float(val or 0.0)) / rate
        return f"${amt:,.2f}"
    except Exception:
        try:
            return f"${float(val):,.2f}"
        except Exception:
            return str(val)

@app.template_filter("money_both")
def money_both(val):
    # e.g. "1,234,567 ل.ل  /  $13.79"
    try:
        return f"{money_lbp(val)}  /  {money_usd(val)}"
    except Exception:
        return str(val)


# Make helpers available in templates
@app.context_processor
def inject_helpers():
    return dict(get_pricing=get_pricing)

def get_pricing() -> Pricing:
    # Robust getter that auto-creates row and self-heals columns if needed
    try:
        p = Pricing.query.order_by(Pricing.id.desc()).first()
    except Exception as e:
        msg = str(e)
        # If columns don't exist yet (SQLite OperationalError e3q8), add them on the fly
        if ("no such column" in msg) or ("has no column named" in msg):
            try:
                rows = db.session.execute(text("PRAGMA table_info(pricing)")).fetchall()
                cols = {r[1] for r in rows}
                if "currency_code" not in cols:
                    db.session.execute(text("ALTER TABLE pricing ADD COLUMN currency_code TEXT NOT NULL DEFAULT 'LBP'"))
                if "usd_rate" not in cols:
                    db.session.execute(text("ALTER TABLE pricing ADD COLUMN usd_rate REAL NOT NULL DEFAULT 90000"))
                db.session.commit()
            except Exception as e2:
                db.session.rollback()
            # Retry after healing
            p = Pricing.query.order_by(Pricing.id.desc()).first()
        else:
            raise

    if not p:
        p = Pricing(unit_price=0.0, fee_20=0.0, fee_15=0.0, fee_10=0.0, fee_5=0.0, currency_code='LBP', usd_rate=90000.0)
        db.session.add(p); db.session.commit()
    return p


def month_bounds(d: date):
    first = d.replace(day=1)
    if first.month == 12:
        next_first = first.replace(year=first.year+1, month=1)
    else:
        next_first = first.replace(month=first.month+1)
    return first, next_first

def last_for_branch(branch_number: str):
    if not branch_number: return None
    return Invoice.query.filter_by(branch_number=branch_number).order_by(Invoice.date.desc(), Invoice.id.desc()).first()

def month_key(d: date) -> str:
    return d.strftime("%Y-%m")

def existing_invoice_for_month(branch_number: str, d: date):
    """Return the most recent invoice for this branch in the same month as date d, or None."""
    if not branch_number or not d:
        return None
    first, next_first = month_bounds(d)
    return (Invoice.query
            .filter(Invoice.branch_number == branch_number)
            .filter(Invoice.date >= first)
            .filter(Invoice.date <  next_first)
            .order_by(Invoice.date.desc(), Invoice.id.desc())
            .first())

def has_invoice_in_month(branch_number: str, d: date) -> bool:
    return existing_invoice_for_month(branch_number, d) is not None

def next_invoice_number_for_date(d: date) -> str:
    yyyymm = d.strftime("%Y%m"); prefix = f"{yyyymm}-"
    rows = Invoice.query.filter(Invoice.invoice_number.like(f"{prefix}%")).all()
    max_suffix = 0
    for r in rows:
        try: max_suffix = max(max_suffix, int(r.invoice_number.split("-")[-1]))
        except Exception: pass
    return f"{prefix}{(max_suffix+1):04d}"

def apply_pricing_defaults(inv: Invoice, pricing: Pricing):
    if inv.unit_price in (None, 0):
        inv.unit_price = float(pricing.unit_price or 0.0)
    if inv.subscription_fee in (None, 0):
        fee = pricing.fee_for_amp(int(inv.subscription_amps or 0))
        inv.subscription_fee = float(fee or 0.0)

# ---------------- Auth ----------------
login_manager = LoginManager(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    try: return db.session.get(User, int(user_id))
    except Exception: return None

def role_required(role):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not getattr(current_user, "is_authenticated", False):
                return redirect(url_for("login"))
            user_role = getattr(current_user, "role", None)
            allowed = (user_role == role) if not isinstance(role, (list,tuple,set)) else (user_role in role)
            if role == "admin" and getattr(current_user, "is_admin", False):
                allowed = True
            if not allowed:
                flash("غير مسموح.", "error"); return redirect(url_for("index", **_current_filter_args()))
            return fn(*args, **kwargs)
        return wrapper
    return decorator

# ---------------- Pricing page ----------------
@app.route("/pricing", methods=["GET","POST"], endpoint="pricing_page")
@login_required
@role_required("admin")

def pricing_page():
    p = get_pricing()
    if request.method == "POST":
        try:
            # Save currency selection and rate
            p.currency_code = (request.form.get("currency_code") or "LBP").upper()
            p.usd_rate = float(request.form.get("usd_rate") or (p.usd_rate or 90000.0))

            # Read values entered in the displayed currency.
            unit_price_in = float(request.form.get("unit_price") or 0)
            fee_20_in     = float(request.form.get("fee_20") or 0)
            fee_15_in     = float(request.form.get("fee_15") or 0)
            fee_10_in     = float(request.form.get("fee_10") or 0)
            fee_5_in      = float(request.form.get("fee_5")  or 0)

            # Store in LBP internally for consistency
            if p.currency_code == "USD":
                rate = float(p.usd_rate or 1.0) or 1.0
                p.unit_price = unit_price_in * rate
                p.fee_20     = fee_20_in * rate
                p.fee_15     = fee_15_in * rate
                p.fee_10     = fee_10_in * rate
                p.fee_5      = fee_5_in * rate
            else:
                p.unit_price = unit_price_in
                p.fee_20     = fee_20_in
                p.fee_15     = fee_15_in
                p.fee_10     = fee_10_in
                p.fee_5      = fee_5_in

            p.updated_at = datetime.utcnow()
            db.session.commit()
            flash("تم حفظ التسعير.", "success")
        except Exception as e:
            db.session.rollback()
            flash("حدث خطأ أثناء الحفظ.", "error")
        return redirect(url_for("pricing_page"))

    # Values to display in the selected currency
    rate = float(p.usd_rate or 1.0) or 1.0
    if p.currency_code == "USD":
        dp = {
            "unit_price": (p.unit_price or 0.0) / rate,
            "fee_20": (p.fee_20 or 0.0) / rate,
            "fee_15": (p.fee_15 or 0.0) / rate,
            "fee_10": (p.fee_10 or 0.0) / rate,
            "fee_5":  (p.fee_5  or 0.0) / rate,
        }
    else:
        dp = {
            "unit_price": p.unit_price or 0.0,
            "fee_20": p.fee_20 or 0.0,
            "fee_15": p.fee_15 or 0.0,
            "fee_10": p.fee_10 or 0.0,
            "fee_5":  p.fee_5  or 0.0,
        }
    return render_template("pricing.html", p=p, dp=dp, now=datetime.utcnow())

# ---------------- Routes ----------------


@app.route("/bulk/update-unit-price", methods=["POST"], endpoint="bulk_update_unit_price")
@login_required
@role_required("admin")
def bulk_update_unit_price():
    """
    Bulk update invoices' unit_price (LBP only) for a specific month/year.
    """
    try:
        month = int(request.form.get("month") or 0)
        year  = int(request.form.get("year") or 0)
        unit_price_input = float(request.form.get("unit_price_input") or 0.0)

        if month < 1 or month > 12 or year < 1900:
            flash("برجاء اختيار شهر وسنة صالحَين.", "error")
            return redirect(request.referrer or url_for("pricing_page"))

        # Always treat input as LBP (no currency conversion)
        unit_price_lbp = unit_price_input

        # Date range for selected month
        start_date = date(year, month, 1)
        end_date = date(year + (1 if month == 12 else 0),
                        1 if month == 12 else month + 1,
                        1)

        # Query and update invoices
        q = Invoice.query.filter(Invoice.date >= start_date, Invoice.date < end_date)
        count = 0
        for inv in q.all():
            inv.unit_price = unit_price_lbp
            inv.kwh_used = max(0, (inv.curr_reading or 0) - (inv.prev_reading or 0))
            inv.energy_cost = round(inv.kwh_used * float(inv.unit_price or 0), 2)
            inv.month_cost = float(inv.subscription_fee or 0)
            inv.total_due = round((inv.energy_cost or 0) + (inv.month_cost or 0), 2)
            count += 1

        db.session.commit()
        flash(f"تم تحديث تسعيرة ك.و.س ({count} فاتورة) لشهر {month:02d}/{year}.", "success")

    except Exception as e:
        db.session.rollback()
        flash("حدث خطأ أثناء التحديث.", "error")

    return redirect(request.referrer or url_for("pricing_page"))
# ---------------- Login ----------------
@app.route("/login", methods=["GET", "POST"], endpoint="login")
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()

        user = User.query.filter_by(username=username).first()
        if not user or not user.check_password(password):
            flash("اسم المستخدم أو كلمة المرور غير صحيحة.", "error")
            return render_template("login.html")

        login_user(user)
        flash("تم تسجيل الدخول بنجاح.", "success")
        return redirect(url_for("index"))

    # GET
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user(); flash("تم تسجيل الخروج.", "success"); return redirect(url_for("login"))

@app.route("/", methods=["GET"])
@login_required
def index():
    # Employee view
    if current_user.role == "employee" and not getattr(current_user, "is_admin", False):
        q = (request.args.get("q") or "").strip().lower()
        ym = (request.args.get("ym") or "").strip()
        status = (request.args.get("status") or "").strip().lower()
        qry = Invoice.query

        # Month filter (YYYY-MM)
        if ym and re.match(r"^\d{4}-\d{2}$", ym):
            y, mth = ym.split("-")
            y = int(y); mth = int(mth)
            from datetime import date as _date
            first = _date(y, mth, 1)
            first, next_first = month_bounds(first)
            qry = qry.filter(Invoice.date >= first, Invoice.date < next_first)

        # Status filter
        if status in ("paid", "unpaid"):
            want_paid = (status == "paid")
            qry = qry.filter(Invoice.is_paid == want_paid)

        rows = qry.order_by(Invoice.id.desc()).all()

        if getattr(current_user, "min_visible_date", None):
            rows = [r for r in rows if r.date is None or r.date >= current_user.min_visible_date]
        if q:
            rows = [r for r in rows if q in (r.branch_number or "").lower() or q in (r.customer_name or "").lower()]
        return render_template("employee_list.html", rows=rows, q=q, this_month=datetime.utcnow().strftime("%Y-%m"), ym=ym, status=status)
        

    # Admin list
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip().lower()
    start_param = (request.args.get("start") or "").strip()
    end_param = (request.args.get("end") or "").strip()

    qry = Invoice.query
    if status == "paid":
        qry = qry.filter(Invoice.is_paid.is_(True))
    elif status == "unpaid":
        qry = qry.filter(Invoice.is_paid.is_(False))

    today = datetime.utcnow().date()
    first, next_first = month_bounds(today)
    start_date = end_date = None
    if not start_param and not end_param:
        start_date = first; end_date = next_first - timedelta(days=1)
        qry = qry.filter(Invoice.date >= start_date).filter(Invoice.date <= end_date)
    else:
        if start_param:
            try: start_date = datetime.strptime(start_param, "%Y-%m-%d").date(); qry = qry.filter(Invoice.date >= start_date)
            except Exception: start_date = None
        if end_param:
            try: end_date = datetime.strptime(end_param, "%Y-%m-%d").date(); qry = qry.filter(Invoice.date <= end_date)
            except Exception: end_date = None

    invoices = qry.order_by(Invoice.id.desc()).all()

    if q:
        ql = q.lower()
        invoices = [i for i in invoices if ql in (i.customer_name or "").lower()
                    or ql in (i.invoice_number or "").lower()
                    or ql in (i.branch_number or "").lower()]

    total_kwh = sum(max(0, (i.curr_reading or 0) - (i.prev_reading or 0)) for i in invoices)
    start_value = (start_date.isoformat() if start_date else (start_param or ""))
    end_value   = (end_date.isoformat()   if end_date   else (end_param   or ""))

    return render_template("list.html", invoices=invoices, q=q, status=status,
                           total_kwh=total_kwh, start=start_value, end=end_value,
                           this_month=datetime.utcnow().strftime("%Y-%m"))

# Employee quick create
@app.post("/employee/quick-create", endpoint="employee_quick_create")
@login_required
@role_required("employee")
def employee_quick_create():
    branch = (request.form.get("branch_number") or "").strip()
    try:
        curr_reading = int(float(request.form.get("curr_reading") or request.form.get("current_reading") or "0"))
    except Exception:
        curr_reading = 0
    last = last_for_branch(branch)
    today = datetime.utcnow().date()
    # prevent duplicate per branch per month
    if has_invoice_in_month(branch, today):
        existing = existing_invoice_for_month(branch, today)
        flash(f"يوجد فاتورة لهذه الشعبة لنفس الشهر ({month_key(today)}): رقم {existing.invoice_number}", "error")
        return redirect(request.referrer or url_for("index"))

    if not last:
        flash("لا توجد فاتورة سابقة لهذه الشعبة.", "error"); return redirect(request.referrer or url_for("index"))
    if curr_reading < (last.curr_reading or 0):
        flash("القراءة الحالية يجب أن تكون ≥ آخر قراءة.", "error"); return redirect(request.referrer or url_for("index"))
    inv = Invoice(invoice_number=next_invoice_number_for_date(today), date=today,
                  customer_name=last.customer_name, meter_number=last.meter_number, branch_number=branch,
                  subscription_amps=last.subscription_amps or 0, prev_reading=last.curr_reading or 0,
                  curr_reading=curr_reading, unit_price=float(get_pricing().unit_price or 0.0), subscription_fee=last.subscription_fee, is_paid=False)
    apply_pricing_defaults(inv, get_pricing())
    inv.kwh_used = max(0, inv.curr_reading - inv.prev_reading)
    inv.energy_cost = round(inv.kwh_used * float(inv.unit_price or 0), 2)
    inv.month_cost = float(inv.subscription_fee or 0)
    inv.total_due = round(inv.energy_cost + inv.month_cost, 2)
    db.session.add(inv); db.session.commit()
    flash("تم إنشاء فاتورة جديدة.", "success"); return redirect(url_for("index", **_current_filter_args()))

# Mark paid / Toggle paid
@app.post("/invoice/<int:invoice_id>/mark-paid", endpoint="mark_paid")
@login_required
@role_required("employee")
def mark_paid(invoice_id: int):
    inv = Invoice.query.get_or_404(invoice_id)
    if inv.is_paid:
        flash("هذه الفاتورة مدفوعة مسبقًا.", "info"); return redirect(request.referrer or url_for("index"))
    inv.is_paid = True; db.session.commit(); flash("تم تعليم الفاتورة كمدفوعة.", "success"); return redirect(request.referrer or url_for("index"))

@app.post("/invoice/<int:invoice_id>/toggle-paid", endpoint="toggle_paid")
@login_required
@role_required("admin")
def toggle_paid(invoice_id: int):
    inv = Invoice.query.get_or_404(invoice_id); inv.is_paid = not bool(inv.is_paid); db.session.commit()
    flash("تم تحديث حالة الدفع.", "success"); return redirect(request.referrer or url_for("index"))

# Export (also mapped to /admin/export)
@app.get("/export", endpoint="export_invoices")
@login_required
@role_required("admin")
def export_invoices():
    fmt = (request.args.get("format") or "csv").lower()
    month = request.args.get("month")
    qry = Invoice.query.order_by(Invoice.id.desc())
    if month:
        try:
            y, m = [int(x) for x in month.split("-")]
            first = date(y, m, 1)
            last_day = (date(y + (m // 12), (m % 12) + 1, 1) - timedelta(days=1))
            qry = qry.filter(Invoice.date >= first, Invoice.date <= last_day)
        except Exception:
            pass
    rows = qry.all()
    headers = ["id","invoice_number","date","customer_name","meter_number","branch_number",
               "subscription_amps","prev_reading","curr_reading","kwh_used","unit_price",
               "energy_cost","subscription_fee","month_cost","total_due","is_paid"]
    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title="invoices"; ws.append(headers)
            for i in rows:
                ws.append([i.id, i.invoice_number, i.date.isoformat() if i.date else "", i.customer_name,
                           i.meter_number, i.branch_number, i.subscription_amps, i.prev_reading, i.curr_reading,
                           i.kwh_used, i.unit_price, i.energy_cost, i.subscription_fee, i.month_cost, i.total_due,
                           int(bool(i.is_paid))])
            bio = io.BytesIO(); wb.save(bio); wb.close(); bio.seek(0)
            return send_file(bio, as_attachment=True, download_name="invoices.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception:
            fmt = "csv"
    sio = io.StringIO(); w = csv.writer(sio); w.writerow(headers)
    for i in rows:
        w.writerow([i.id, i.invoice_number, i.date.isoformat() if i.date else "", i.customer_name,
                    i.meter_number, i.branch_number, i.subscription_amps, i.prev_reading, i.curr_reading,
                    i.kwh_used, i.unit_price, i.energy_cost, i.subscription_fee, i.month_cost, i.total_due,
                    int(bool(i.is_paid))])
    return Response(sio.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition":"attachment; filename=invoices.csv"})

try:
    app.add_url_rule("/admin/export", endpoint="export_invoices", view_func=export_invoices, methods=["GET"])
except Exception:
    pass

# Print & PDF


@app.get("/invoices/print-all", endpoint="print_all_invoices")
@login_required
@role_required("admin")
def print_all_invoices():
    """
    Render a printable view that includes all invoices for the specified month.
    Accepts ?month=YYYY-MM (preferred). If not provided, will derive from ?start=YYYY-MM-DD.
    """
    month_param = (request.args.get("month") or "").strip()
    start_param = (request.args.get("start") or "").strip()
    end_param   = (request.args.get("end") or "").strip()

    # Determine month key
    month_key_str = None
    if re.match(r"^\d{4}-\d{2}$", month_param or ""):
        month_key_str = month_param
        year, mon = map(int, month_param.split("-"))
        first_day = date(year, mon, 1)
    elif re.match(r"^\d{4}-\d{2}-\d{2}$", start_param or ""):
        y, m, d = map(int, start_param.split("-"))
        first_day = date(y, m, 1)
        month_key_str = first_day.strftime("%Y-%m")
    else:
        # default current UTC month
        first_day = datetime.utcnow().date().replace(day=1)
        month_key_str = first_day.strftime("%Y-%m")

    # Compute end exclusive
    _, next_first = month_bounds(first_day)

    # Query invoices whose date within [first_day, next_first)
    rows = (Invoice.query
            .filter(Invoice.date >= first_day)
            .filter(Invoice.date <  next_first)
            .order_by(Invoice.id.asc())
            .all())

    # Precompute QR codes if available (optional, same as invoice_print)
    qr_map = {}
    if qrcode:
        for i in rows:
            try:
                qr = qrcode.QRCode(version=1, box_size=2, border=1)
                qr.add_data(i.invoice_number or "")
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                bio = io.BytesIO(); img.save(bio, format="PNG")
                qr_map[i.id] = "data:image/png;base64," + base64.b64encode(bio.getvalue()).decode("ascii")
            except Exception:
                qr_map[i.id] = None

    return render_template("invoice_print_all.html", invoices=rows, month=month_key_str, qr_map=qr_map)
@app.get("/invoice/<int:invoice_id>/print", endpoint="invoice_print")
@login_required
@role_required("admin")
def invoice_print(invoice_id: int):
    i = Invoice.query.get_or_404(invoice_id)
    qr_data_uri = None
    if qrcode:
        try:
            qr = qrcode.QRCode(version=1, box_size=4, border=1)
            qr.add_data(f"INV:{i.invoice_number}")
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            bio=io.BytesIO(); img.save(bio, format="PNG")
            qr_data_uri = "data:image/png;base64," + base64.b64encode(bio.getvalue()).decode("ascii")
        except Exception:
            qr_data_uri=None
    return render_template("invoice_print.html", i=i, qr_data_uri=qr_data_uri)

@app.get("/invoice/<int:invoice_id>/pdf", endpoint="invoice_pdf")
@login_required
@role_required("admin")
def invoice_pdf(invoice_id: int):
    i = Invoice.query.get_or_404(invoice_id)
    html = render_template("invoice_print.html", i=i, qr_data_uri=None)
    try:
        from xhtml2pdf import pisa
        pdf_io = io.BytesIO()
        status = pisa.CreatePDF(html, dest=pdf_io, encoding="utf-8")
        if not status.err:
            pdf_io.seek(0)
            return Response(pdf_io.getvalue(), mimetype="application/pdf",
                            headers={"Content-Disposition": f"attachment; filename=invoice_{invoice_id}.pdf"})
    except Exception:
        pass
    flash("تعذّر إنشاء PDF تلقائيًا. استخدم الطباعة ثم حفظ كـ PDF.", "error")
    return redirect(url_for("invoice_print", invoice_id=invoice_id))

# Users (minimal)


@app.route("/admin/users", methods=["GET","POST"], endpoint="manage_users")
@login_required
@role_required("admin")
def manage_users():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        role = (request.form.get("role") or "employee").strip()
        branch_number = (request.form.get("branch_number") or "").strip() or None
        is_admin = bool(request.form.get("is_admin"))
        min_visible_date_str = (request.form.get("min_visible_date") or "").strip()
        min_visible_date = None
        if min_visible_date_str:
            try:
                min_visible_date = datetime.strptime(min_visible_date_str, "%Y-%m-%d").date()
            except Exception:
                flash("تاريخ غير صالح. الرجاء استخدام الصيغة YYYY-MM-DD", "danger")
                return redirect(url_for("manage_users"))

        if not username or not password:
            flash("الرجاء إدخال اسم المستخدم وكلمة المرور.", "danger")
            return redirect(url_for("manage_users"))

        if User.query.filter_by(username=username).first():
            flash("اسم المستخدم مستخدم بالفعل.", "danger")
            return redirect(url_for("manage_users"))

        u = User(username=username, role=role, branch_number=branch_number, min_visible_date=min_visible_date, is_admin=is_admin)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash("تم إنشاء المستخدم بنجاح.", "success")
        return redirect(url_for("manage_users"))

    users = User.query.order_by(User.id.desc()).all()
    return render_template("users.html", users=users)
@app.route("/admin/users/<int:user_id>/edit", methods=["GET","POST"], endpoint="edit_user")
@login_required
@role_required("admin")
def edit_user(user_id):
    u = User.query.get_or_404(user_id)
    if request.method == "POST":
        new_username = (request.form.get("username") or "").strip()
        new_password = (request.form.get("password") or "").strip()
        new_role = (request.form.get("role") or "").strip()
        branch_number = (request.form.get("branch_number") or "").strip() or None
        is_admin = bool(request.form.get("is_admin"))
        min_visible_date_str = (request.form.get("min_visible_date") or "").strip()
        min_visible_date = None
        if min_visible_date_str:
            try:
                min_visible_date = datetime.strptime(min_visible_date_str, "%Y-%m-%d").date()
            except Exception:
                flash("تاريخ غير صالح.", "danger")
                return redirect(url_for("edit_user", user_id=user_id))

        if new_username:
            # ensure unique
            existing = User.query.filter(User.username == new_username, User.id != u.id).first()
            if existing:
                flash("اسم المستخدم مستخدم بالفعل.", "danger")
                return redirect(url_for("edit_user", user_id=user_id))
            u.username = new_username
        if new_password:
            u.set_password(new_password)
        if new_role:
            u.role = new_role
        u.branch_number = branch_number
        u.is_admin = is_admin
        u.min_visible_date = min_visible_date

        db.session.commit()
        flash("تم تحديث المستخدم.", "success")
        return redirect(url_for("manage_users"))
    return render_template("user_edit.html", user=u, u=u)
def manage_users():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        role = (request.form.get("role") or "employee").strip()
        branch_number = (request.form.get("branch_number") or "").strip() or None
        is_admin = bool(request.form.get("is_admin"))
        min_visible_date_str = (request.form.get("min_visible_date") or "").strip()
        min_visible_date = None
        if min_visible_date_str:
            try:
                min_visible_date = datetime.strptime(min_visible_date_str, "%Y-%m-%d").date()
            except Exception:
                flash("تاريخ غير صالح. الرجاء استخدام الصيغة YYYY-MM-DD", "danger")
                return redirect(url_for("manage_users"))

        if not username or not password:
            flash("الرجاء إدخال اسم المستخدم وكلمة المرور.", "danger")
            return redirect(url_for("manage_users"))

        # Check if username exists
        existing = User.query.filter_by(username=username).first()
        if existing:
            flash("اسم المستخدم مستخدم بالفعل.", "danger")
            return redirect(url_for("manage_users"))

        # Create user
        u = User(username=username, role=role, branch_number=branch_number, min_visible_date=min_visible_date, is_admin=is_admin)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash("تم إنشاء المستخدم بنجاح.", "success")
        return redirect(url_for("manage_users"))

    users = User.query.order_by(User.id.desc()).all()
    return render_template("users.html", users=users)


# New / View / Edit / Delete Invoice


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"], endpoint="delete_user")
@login_required
@role_required("admin")
def delete_user(user_id):
    u = User.query.get_or_404(user_id)
    # Prevent deleting your own account accidentally
    if current_user.id == u.id:
        flash("لا يمكنك حذف حسابك أثناء تسجيل الدخول.", "warning")
        return redirect(url_for("manage_users"))
    db.session.delete(u)
    db.session.commit()
    flash("تم حذف المستخدم.", "success")
    return redirect(url_for("manage_users"))
@app.route("/invoice/new", methods=["GET","POST"], endpoint="new_invoice")
@login_required
@role_required("admin")
def new_invoice():
    if request.method == "POST":
        # 1) read date from the form (التاريخ)
        date_str = (request.form.get("date") or "").strip()

        def _parse_date(s: str):
            # supports HTML date input and common day/month formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            return None

        use_date = _parse_date(date_str) or datetime.utcnow().date()

        # 2) read other fields
        branch = (request.form.get("branch_number") or "").strip()
        customer_name = (request.form.get("customer_name") or "").strip()
        meter_number = (request.form.get("meter_number") or "").strip()

        try:
            subscription_amps = int(request.form.get("subscription_amps") or 0)
        except Exception:
            subscription_amps = 0

        # defaults from last invoice of the same branch
        last = last_for_branch(branch)

        # 3) prevent duplicate for the *selected* month (not today's)
        if has_invoice_in_month(branch, use_date):
            existing = existing_invoice_for_month(branch, use_date)
            flash(
                f"يوجد فاتورة لهذه الشعبة لنفس الشهر ({month_key(use_date)}): رقم {existing.invoice_number}",
                "error",
            )
            return redirect(request.referrer or url_for("index"))

        # 4) readings
        try:
            prev_reading = int(request.form.get("prev_reading") or (getattr(last, "curr_reading", 0) or 0))
        except Exception:
            prev_reading = int(getattr(last, "curr_reading", 0) or 0)
        try:
            curr_reading = int(request.form.get("curr_reading") or prev_reading)
        except Exception:
            curr_reading = prev_reading

        if not branch:
            flash("رقم الشعبة مطلوب.", "error")
            return redirect(request.referrer or url_for("new_invoice"))

        if not customer_name:
            customer_name = getattr(last, "customer_name", "") or customer_name

        def _float(v):
            try:
                return float(v) if (v is not None and v != "") else None
            except Exception:
                return None

        unit_price = _float(request.form.get("unit_price"))
        if unit_price is None:
            unit_price = float(get_pricing().unit_price or 0.0)
        subscription_fee = _float(request.form.get("subscription_fee"))

        # 5) create invoice using the selected date
        inv = Invoice(
            invoice_number=next_invoice_number_for_date(use_date),
            date=use_date,
            customer_name=customer_name,
            meter_number=meter_number or getattr(last, "meter_number", "") or "",
            branch_number=branch,
            subscription_amps=subscription_amps or (getattr(last, "subscription_amps", 0) or 0),
            prev_reading=prev_reading,
            curr_reading=curr_reading,
            unit_price = (unit_price if unit_price is not None else float(get_pricing().unit_price or 0.0)),
            subscription_fee=subscription_fee if subscription_fee is not None else (getattr(last, "subscription_fee", None) or get_pricing().fee_for_amp(subscription_amps or (getattr(last, "subscription_amps", 0) or 0))),
            is_paid=is_paid_val,
        )

        # 6) totals
        apply_pricing_defaults(inv, get_pricing())
        inv.kwh_used = max(0, inv.curr_reading - inv.prev_reading)
        inv.energy_cost = round(inv.kwh_used * float(inv.unit_price or 0), 2)
        inv.month_cost = float(inv.subscription_fee or 0)
        inv.total_due = round(inv.energy_cost + inv.month_cost, 2)

        db.session.add(inv)
        db.session.commit()
        flash("تم إنشاء الفاتورة الجديدة.", "success")
        return redirect(url_for("index", **_current_filter_args()))

    # ----- GET: preview defaults using ?branch= -----
    branch_arg = (request.args.get("branch") or "").strip()
    last = last_for_branch(branch_arg) if branch_arg else None
    default_unit = float(get_pricing().unit_price or 0.0)
    today = date.today().strftime("%Y-%m-%d")
    return render_template("invoice_new.html", last=last, branch=branch_arg, default_unit=default_unit, today=today)

@app.get("/invoice/<int:invoice_id>", endpoint="view_invoice")
@login_required
@role_required("admin")
def view_invoice(invoice_id: int):
    i = Invoice.query.get_or_404(invoice_id)
    return render_template("invoice_view.html", i=i)

# ---- Edit / Delete (admin) ----
@app.route("/invoice/<int:invoice_id>/edit", methods=["GET", "POST"], endpoint="edit_invoice")
@login_required
@role_required("admin")
def edit_invoice(invoice_id: int):
    i = Invoice.query.get_or_404(invoice_id)

    if request.method == "POST":
        # helpers
        def _date(v, default=i.date):
            try:
                return datetime.strptime((v or "").strip(), "%Y-%m-%d").date()
            except Exception:
                return default

        def _int(v, default=0):
            try: return int(float(v))
            except Exception: return default

        def _float(v, default=None):
            try:
                return float(v) if v not in (None, "") else default
            except Exception:
                return default

        # read form
        new_date   = _date(request.form.get("date"))
        new_branch = (request.form.get("branch_number") or i.branch_number or "").strip()
        new_name   = (request.form.get("customer_name") or i.customer_name or "").strip()
        new_meter  = (request.form.get("meter_number") or i.meter_number or "").strip()
        new_amps   = _int(request.form.get("subscription_amps"), i.subscription_amps or 0)
        new_prev   = _int(request.form.get("prev_reading"), i.prev_reading or 0)
        new_curr   = _int(request.form.get("curr_reading"), i.curr_reading or 0)
        new_unit   = _float(request.form.get("unit_price"), i.unit_price)
        new_fee    = _float(request.form.get("subscription_fee"), i.subscription_fee)
        new_paid   = request.form.get("is_paid") in ("on", "1", "true", "yes")

        # duplicate guard (same شعبة + same month, excluding this invoice)
        first, next_first = month_bounds(new_date)
        dup = (Invoice.query
               .filter(Invoice.id != i.id)
               .filter(Invoice.branch_number == new_branch)
               .filter(Invoice.date >= first, Invoice.date < next_first)
               .first())
        if dup:
            flash(f"يوجد فاتورة لهذه الشعبة لنفس الشهر ({new_date.strftime('%Y-%m')}): رقم {dup.invoice_number}", "error")
            return redirect(request.referrer or url_for("index"))

        # apply changes
        i.date = new_date
        i.branch_number = new_branch
        i.customer_name = new_name
        i.meter_number = new_meter
        i.subscription_amps = new_amps
        i.prev_reading = new_prev
        i.curr_reading = new_curr
        i.unit_price = new_unit if new_unit is not None else i.unit_price
        i.subscription_fee = new_fee if new_fee is not None else i.subscription_fee
        i.is_paid = new_paid

        # recompute totals
        i.kwh_used   = max(0, i.curr_reading - i.prev_reading)
        i.energy_cost = round(i.kwh_used * float(i.unit_price or 0), 2)
        i.month_cost  = float(i.subscription_fee or 0)
        i.total_due   = round(i.energy_cost + i.month_cost, 2)

        db.session.commit()
        flash("تم تحديث الفاتورة.", "success")
        return redirect(url_for("view_invoice", invoice_id=i.id))

    return render_template("invoice_edit.html", i=i)

@app.post("/invoice/<int:invoice_id>/delete", endpoint="delete_invoice")
@login_required
@role_required("admin")
def delete_invoice(invoice_id: int):
    i = Invoice.query.get_or_404(invoice_id)
    db.session.delete(i)
    db.session.commit()
    flash("تم حذف الفاتورة.", "success")
    return redirect(url_for("index", **_current_filter_args()))


@app.post("/invoices/bulk-delete", endpoint="bulk_delete_invoices")
@login_required
@role_required("admin")
def bulk_delete_invoices():
    ids = request.form.getlist("invoice_ids")
    if not ids:
        ids = request.form.getlist("invoice_ids[]")
    cleaned = []
    for _id in ids:
        try:
            cleaned.append(int(_id))
        except Exception:
            continue
    if not cleaned:
        flash("لم يتم تحديد أي فواتير.", "warning")
        return redirect(url_for("index", **_current_filter_args()))
    try:
        Invoice.query.filter(Invoice.id.in_(cleaned)).delete(synchronize_session=False)
        db.session.commit()
        flash(f"تم حذف {len(cleaned)} فاتورة.", "success")
    except Exception as e:
        db.session.rollback()
        flash("حصل خطأ أثناء حذف الفواتير.", "error")
    return redirect(url_for("index", **_current_filter_args()))

# Import branches + template
@app.get("/admin/branches/template.xlsx", endpoint="branches_template_xlsx")
@login_required
@role_required("admin")
def branches_template_xlsx():
    try:
        from openpyxl import Workbook
    except Exception:
        flash("الرجاء تثبيت openpyxl لإنشاء القالب: pip install openpyxl", "error")
        return redirect(request.referrer or url_for("index"))
    wb = Workbook(); ws = wb.active; ws.title="branches"
    # أضفنا الأعمدة الجديدة: prev_reading, invoice_date, invoice_number
    ws.append([
        "branch_number","customer_name","meter_number","subscription_amps",
        "unit_price","subscription_fee","curr_reading",
        "prev_reading","invoice_date","invoice_number","is_paid"
    ])
    # مثال صف عيّنة
    ws.append(["101","أحمد محمد","M-1001",10,0.12,20000,3500,3200,"2025-09-01","202509-0001",0])
    bio = io.BytesIO(); wb.save(bio); wb.close(); bio.seek(0)
    return send_file(
        bio, as_attachment=True, download_name="branches_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/branches/import-xlsx", methods=["GET","POST"], endpoint="import_branches_xlsx")
@login_required
@role_required("admin")
def import_branches_xlsx():
    import io, csv
    from datetime import datetime, date

    if request.method == "GET":
        return render_template("import_branches.html", excel=True)

    f = request.files.get("file")
    if not f or not f.filename:
        flash("اختر ملف Excel أو CSV.", "error")
        return redirect(url_for("import_branches_xlsx"))
    name = f.filename.lower()
    today = datetime.utcnow().date()
    created = 0

    # تحويل تاريخ Excel serial أو نص إلى date
    def _parse_date(val):
        if val in (None, ""):
            return None
        # Excel serial numbers (1900-based), handle without pandas
        try:
            if isinstance(val, (int, float)) and float(val) > 10000:
                base = datetime(1899, 12, 30)
                return (base + timedelta(days=int(val))).date()
        except Exception:
            pass
        s = str(val).strip()
        # Try multiple common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        # Last resort: split heuristics (m/d/y or y-m)
        try:
            parts = re.split(r"[\-/]", s)
            if len(parts) == 3:
                a,b,c = parts
                if len(a)==4:  # Y-M-D
                    return date(int(a), int(b), int(c))
                # assume M/D/Y
                return date(int(c), int(a), int(b))
        except Exception:
            return None

    def create_from_row(getv):
        nonlocal created
        branch = str(getv("branch_number") or "").strip()
        if not branch:
            return

        # آخر فاتورة للشعبة لافتراض القراءة السابقة إذا ما أُرسلت
        last = last_for_branch(branch)

        # القيم الأساسية (مع fallbacks)
        customer_name = (str(getv("customer_name") or "") or getattr(last, "customer_name", "")).strip() or getattr(last, "customer_name", "")
        meter_number  = (str(getv("meter_number")  or "") or getattr(last, "meter_number",  "")).strip() or getattr(last, "meter_number",  "")

        try:
            subscription_amps = int(getv("subscription_amps") or (getattr(last, "subscription_amps", 0) or 0))
        except Exception:
            subscription_amps = getattr(last, "subscription_amps", 0) or 0

        try:
            unit_price = float(getv("unit_price") or (getattr(last, "unit_price", 0.0) or 0.0))
        except Exception:
            unit_price = getattr(last, "unit_price", 0.0) or 0.0

        try:
            subscription_fee = float(getv("subscription_fee") or (getattr(last, "subscription_fee", 0.0) or 0.0))
        except Exception:
            subscription_fee = getattr(last, "subscription_fee", 0.0) or 0.0

        # القيم الجديدة (مؤثرة على إنشاء الفاتورة)
        # prev_reading إن لم يرد بالملف سنأخذ من آخر فاتورة
        prev_from_file = getv("prev_reading")
        try:
            prev = int(float(prev_from_file)) if prev_from_file not in (None, "") else int(getattr(last, "curr_reading", 0) or 0)
        except Exception:
            prev = int(getattr(last, "curr_reading", 0) or 0)

        # current reading
        try:
            curr = int(float(getv("curr_reading") or prev))
        except Exception:
            curr = prev

        # تاريخ الفاتورة، إن لم يُرسل بالملف نستخدم اليوم
        inv_date_from_file = _parse_date(getv("invoice_date"))
        inv_date = inv_date_from_file or today

        # رقم الفاتورة: نستخدم المرسل إن وُجد، وإلا نتولّد بحسب التاريخ
        inv_number_from_file = (str(getv("invoice_number") or "").strip() or None)
        invoice_number = inv_number_from_file or next_invoice_number_for_date(inv_date)

        # حالة الدفع من الملف (0/1, true/false, نعم/لا)
        _is_paid_raw = getv("is_paid")
        def _to_bool(v):
            if v is None: return False
            s = str(v).strip().lower()
            return s in ("1","true","yes","y","نعم","مدفوع")
        is_paid_val = _to_bool(_is_paid_raw)


        inv = Invoice(
            invoice_number=invoice_number,
            date=inv_date,
            customer_name=customer_name,
            meter_number=meter_number,
            branch_number=branch,
            subscription_amps=subscription_amps,
            prev_reading=prev,
            curr_reading=curr,
            unit_price=unit_price,
            subscription_fee=subscription_fee,
            is_paid=is_paid_val,
        )

        # تسعير افتراضي/قواعد إضافية
        apply_pricing_defaults(inv, get_pricing())

        # حسابات
        inv.kwh_used    = max(0, curr - prev)
        inv.energy_cost = round(inv.kwh_used * float(inv.unit_price or 0), 2)
        inv.month_cost  = float(inv.subscription_fee or 0)
        inv.total_due   = round(inv.energy_cost + inv.month_cost, 2)

        db.session.add(inv); created += 1

    def _norm(s): return (str(s or "")).strip().lower().replace(" ", "").replace("_","")

    # أضفنا مرادفات للأعمدة الجديدة
    aliases = {
        "branch_number"    : ["branch_number","branch","الشعبة","رقمالشعبة","شعبة"],
        "customer_name"    : ["customer_name","customer","الاسم","اسم","المشترك"],
        "meter_number"     : ["meter_number","meter","العداد","رقمالعداد"],
        "subscription_amps": ["subscription_amps","amps","الأمبير","الامبير"],
        "unit_price"       : ["unit_price","سعرالوحدة","سعر_الوحدة","سعرالكيلوواط","سعركيلوواط"],
        "subscription_fee" : ["subscription_fee","اشتراكمشهري","رسماشتراك","اشتراك"],
        "curr_reading"     : ["curr_reading","current_reading","القراءةالحالية","قراءةحالية","قراءة"],
        # الجدد:
        "prev_reading"     : ["prev_reading","previous_reading","السابق","القراءةالسابقة","قراءةسابقة"],
        "invoice_date"     : ["invoice_date","date","تاريخ","تاريخ_الفاتورة","تاريخالفاتورة"],
        "invoice_number"   : ["invoice_number","invoice_no","no","رقمالفاتورة","رقم_الفاتورة"],
    
        "is_paid"          : ["is_paid","paid","status","مدفوع","حالةالدفع","حالة_الدفع"]
    }

    def build_map(header_row):
        idx={}; norm=[_norm(h) for h in header_row]
        for canon,names in aliases.items():
            for n in names:
                if _norm(n) in norm:
                    idx[canon]=norm.index(_norm(n)); break
        return idx

    if name.endswith((".xlsx",".xls")):
        try:
            from openpyxl import load_workbook
        except Exception:
            flash("يلزم تثبيت openpyxl لاستيراد Excel: pip install openpyxl", "error")
            return redirect(url_for("import_branches_xlsx"))

        wb = load_workbook(f, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None) or []
        hmap = build_map(header) if header else {}

        # إن ما عرفنا الرأس، نستعمل ترتيب افتراضي يشمل الأعمدة الجديدة أيضًا (اختيارية آخر 3)
        if not hmap:
            headers = [
                "branch_number","customer_name","meter_number","subscription_amps",
                "unit_price","subscription_fee","curr_reading",
                "prev_reading","invoice_date","invoice_number","is_paid"
            ]
            hmap = {k:i for i,k in enumerate(headers)}
            rows = ws.iter_rows(min_row=1, values_only=True)

        for r in rows:
            def getv(k):
                idx=hmap.get(k)
                return r[idx] if (idx is not None and isinstance(r,(list,tuple)) and idx < len(r)) else None
            create_from_row(getv)

        wb.close()
        db.session.commit()

    elif name.endswith(".csv"):
        raw = f.read()
        try:
            text_data = raw.decode("utf-8-sig")
        except Exception:
            try:
                text_data = raw.decode("utf-8")
            except Exception:
                text_data = raw.decode("cp1256", errors="ignore")
        rdr = csv.DictReader(io.StringIO(text_data))

        if rdr.fieldnames:
            hmap = build_map(rdr.fieldnames)
            if hmap:
                for row in rdr:
                    def getv(k):
                        # ابحث بأي مرادف
                        for alias in aliases.get(k, []):
                            for key in row.keys():
                                if _norm(key) == _norm(alias):
                                    return row.get(key)
                        return None
                    create_from_row(getv)
            else:
                # بدون رؤوس معروفة: ترتيب افتراضي
                f2 = csv.reader(io.StringIO(text_data))
                for row in f2:
                    order = [
                        "branch_number","customer_name","meter_number","subscription_amps",
                        "unit_price","subscription_fee","curr_reading",
                        "prev_reading","invoice_date","invoice_number"
                    ]
                    def getv(k):
                        idx=order.index(k) if k in order else None
                        return row[idx] if idx is not None and idx < len(row) else None
                    create_from_row(getv)
        else:
            # ملف CSV بلا رؤوس
            f2 = csv.reader(io.StringIO(text_data))
            for row in f2:
                order = [
                    "branch_number","customer_name","meter_number","subscription_amps",
                    "unit_price","subscription_fee","curr_reading",
                    "prev_reading","invoice_date","invoice_number"
                ]
                def getv(k):
                    idx=order.index(k) if k in order else None
                    return row[idx] if idx is not None and idx < len(row) else None
                create_from_row(getv)

        db.session.commit()

    else:
        flash("صيغة الملف غير مدعومة.", "error")
        return redirect(url_for("import_branches_xlsx"))

    flash(f"تم استيراد {created} صفًا.", "success")
    return redirect(url_for("index", **_current_filter_args()))


# Report
@app.get("/report", endpoint="report")
@login_required
@role_required("admin")
def report():
    q_start = (request.args.get("start") or "").strip()
    q_end = (request.args.get("end") or "").strip()
    today = datetime.utcnow().date()
    first, next_first = month_bounds(today)
    start_date = end_date = None
    if not q_start and not q_end:
        start_date = first; end_date = next_first - timedelta(days=1)
    else:
        if q_start:
            try: start_date = datetime.strptime(q_start, "%Y-%m-%d").date()
            except Exception: start_date=None
        if q_end:
            try: end_date = datetime.strptime(q_end, "%Y-%m-%d").date()
            except Exception: end_date=None
    qry = Invoice.query
    if start_date: qry = qry.filter(Invoice.date >= start_date)
    if end_date:   qry = qry.filter(Invoice.date <= end_date)
    rows = qry.order_by(Invoice.date.asc(), Invoice.id.asc()).all()
    total_invoices = len(rows)
    total_kwh = sum(int(getattr(r,"kwh_used",(r.curr_reading or 0)-(r.prev_reading or 0)) or 0) for r in rows)
    total_amount = round(sum(float(r.total_due or 0) for r in rows), 2)
    paid_count = sum(1 for r in rows if r.is_paid)
    unpaid_count = total_invoices - paid_count
    return render_template("report.html", rows=rows, total_invoices=total_invoices, total_kwh=total_kwh,
                           total_amount=total_amount, paid_count=paid_count, unpaid_count=unpaid_count,
                           start=(start_date.isoformat() if start_date else (q_start or "")),
                           end=(end_date.isoformat() if end_date else (q_end or "")))

# API: latest pricing ----
@app.get("/api/pricing/latest", endpoint="api_pricing_latest")
@login_required
def api_pricing_latest():
    p = Pricing.query.order_by(Pricing.updated_at.desc()).first()
    if not p:
        return jsonify({"unit_price": 0.0, "fees": {"5":0.0,"10":0.0,"15":0.0,"20":0.0}})
    return jsonify({
        "unit_price": float(p.unit_price or 0.0),
        "fees": {
            "5":  float(p.fee_5  or 0.0),
            "10": float(p.fee_10 or 0.0),
            "15": float(p.fee_15 or 0.0),
            "20": float(p.fee_20 or 0.0)
        }
    })
# === Expenses ===
class Expense(db.Model):
    __tablename__ = "expenses"
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    type = db.Column(db.String(32), nullable=False)  # 'fuel', 'maintenance', 'other'
    cost = db.Column(db.Float, nullable=False, default=0.0)
    litres = db.Column(db.Float, nullable=True)      # only for fuel
    description = db.Column(db.String(500), nullable=True)

    def to_dict(self):
        return {
            "id": self.id,
            "date": self.date.isoformat(),
            "type": self.type,
            "cost": float(self.cost or 0),
            "litres": float(self.litres or 0) if self.litres is not None else None,
            "description": self.description or ""
        }



@app.route("/expenses", methods=["GET", "POST"], endpoint="expenses")
@login_required
def expenses():
    if request.method == "POST":
        t = (request.form.get("type") or "").strip()
        cost = float(request.form.get("cost") or 0)
        litres = request.form.get("litres")
        litres = float(litres) if litres not in (None, "",) else None
        desc = (request.form.get("description") or "").strip()
        d = request.form.get("date")
        try:
            d = datetime.strptime(d, "%Y-%m-%d").date() if d else datetime.utcnow().date()
        except Exception:
            d = datetime.utcnow().date()
        if t not in ("fuel", "maintenance", "other"):
            flash("نوع المصروف غير صالح.", "error")
        else:
            e = Expense(type=t, cost=cost, litres=litres, description=desc, date=d)
            db.session.add(e); db.session.commit()
            flash("تمت إضافة المصروف.", "success")
        return redirect(url_for("expenses"))
    # GET
    q = Expense.query.order_by(Expense.date.desc(), Expense.id.desc()).limit(500).all()
    total = sum(e.cost or 0 for e in q)
    return render_template("expenses.html", expenses=q, total=total)


@app.route("/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
def delete_expense(expense_id):
    e = Expense.query.get_or_404(expense_id)
    db.session.delete(e); db.session.commit()
    flash("تم حذف المصروف.", "success")
    return redirect(url_for("expenses"))


@app.route("/api/expenses/summary")
@login_required
def api_expenses_summary():
    # Optional date filters
    start = request.args.get("start")
    end = request.args.get("end")
    qry = Expense.query
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qry = qry.filter(Expense.date >= sd)
        except Exception:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qry = qry.filter(Expense.date <= ed)
        except Exception:
            pass
    data = [e.to_dict() for e in qry.all()]
    # Aggregate by type and by date
    by_type = {"fuel": 0.0, "maintenance": 0.0, "other": 0.0}
    by_day = {}
    fuel_litres = 0.0
    for d in data:
        by_type[d["type"]] = by_type.get(d["type"], 0.0) + float(d["cost"] or 0)
        day = d["date"][:10]
        by_day[day] = by_day.get(day, 0.0) + float(d["cost"] or 0)
        if d["type"] == "fuel" and d.get("litres") is not None:
            fuel_litres += float(d["litres"] or 0)
    return jsonify({
        "items": data,
        "by_type": by_type,
        "by_day": [{"date": k, "total": v} for k, v in sorted(by_day.items())],
        "fuel_litres": fuel_litres
    })



# ---- DB bootstrap & lightweight migration ----
with app.app_context():
    try:
        db.create_all()
    except Exception:
        pass
    # users.is_admin (one-time)
    try:
        rows = db.session.execute(text("PRAGMA table_info(users)")).fetchall()
        cols = {r[1] for r in rows}
        if "is_admin" not in cols:
            db.session.execute(text("ALTER TABLE users ADD COLUMN is_admin BOOLEAN NOT NULL DEFAULT 0"))
            db.session.commit()
            print("[migrate] Added users.is_admin")
    except Exception as e:
        print("[migrate] users.is_admin check:", e)
    # pricing.fee_15 (one-time)
    try:
        rows = db.session.execute(text("PRAGMA table_info(pricing)")).fetchall()
        cols = {r[1] for r in rows}
        if "fee_15" not in cols:
            db.session.execute(text("ALTER TABLE pricing ADD COLUMN fee_15 REAL NOT NULL DEFAULT 0"))
            db.session.commit()
            print("[migrate] Added pricing.fee_15")
    except Exception as e:
        print("[migrate] pricing.fee_15 check:", e)



@app.route('/dashboards')
@login_required
def dashboards():
    # ----- inputs: start=end=YYYY-MM -----
    start_str = (request.args.get('start') or '').strip()
    end_str   = (request.args.get('end') or '').strip()

    def ym_to_date(s):
        try:
            y, m = s.split('-')
            from datetime import date
            return date(int(y), int(m), 1)
        except Exception:
            return None

    start_date = ym_to_date(start_str)
    end_month  = ym_to_date(end_str)

    # upper bound (first day of month after end)
    next_after_end = None
    if end_month:
        if end_month.month == 12:
            next_after_end = end_month.replace(year=end_month.year+1, month=1)
        else:
            next_after_end = end_month.replace(month=end_month.month+1)

    # ----- USD rate (LBP per USD) -----
    p = get_pricing()
    rate = float(p.usd_rate or 1.0) or 1.0

    # ----- Invoices aggregated by month (LBP in DB) -----
    qry = (db.session.query(
        func.strftime('%Y-%m', Invoice.date).label('ym'),
        func.count(Invoice.id).label('count'),
        func.sum(Invoice.total_due).label('total_due'),
        func.sum(Invoice.kwh_used).label('kwh'),
        func.sum(case((Invoice.is_paid == True, Invoice.total_due), else_=0.0)).label('paid_due')
    ))
    if start_date:
        qry = qry.filter(Invoice.date >= start_date)
    if next_after_end:
        qry = qry.filter(Invoice.date < next_after_end)

    rows = qry.group_by('ym').order_by('ym').all()
    labels = [r.ym for r in rows]

    # if no data, return safe empty payload
    if not rows:
        return render_template('dashboards.html',
                               labels=[], invoice_counts=[], totals=[],
                               kwh=[], paid=[], unpaid=[], avg_invoice=[],
                               latest={"month":"", "count":0, "total":0.0, "kwh":0,
                                       "paid":0.0, "unpaid":0.0, "avg_invoice":0.0,
                                       "expenses":0.0, "net_total":0.0},
                               start=start_str, end=end_str)

    # chosen month = user selected (start) else last month with data
    selected_ym = start_str if start_str else labels[-1]
    try:
        sel_idx = labels.index(selected_ym)
    except ValueError:
        sel_idx = len(labels) - 1

    # ----- Convert invoice money to USD (once) -----
    invoice_counts = [int(r.count or 0) for r in rows]
    kwh            = [int(r.kwh or 0) for r in rows]

    totals_usd = [ (float(r.total_due or 0.0) / rate) for r in rows ]
    paid_usd   = [ (float(r.paid_due  or 0.0) / rate) for r in rows ]
    unpaid_usd = [ max(0.0, (float(r.total_due or 0.0) - float(r.paid_due or 0.0)) / rate) for r in rows ]
    avg_inv_usd= [ (totals_usd[i] / invoice_counts[i]) if invoice_counts[i] else 0.0
                   for i in range(len(rows)) ]

    # ----- Expenses aggregated by month (ALREADY USD — NO division) -----
    exp_q = db.session.query(
        func.strftime('%Y-%m', Expense.date).label('ym'),
        func.sum(Expense.cost).label('exp_total')
    )
    if start_date:
        exp_q = exp_q.filter(Expense.date >= start_date)
    if next_after_end:
        exp_q = exp_q.filter(Expense.date < next_after_end)
    exp_rows = exp_q.group_by('ym').order_by('ym').all()

    # keep as-is in USD
    exp_map = { r.ym: float(r.exp_total or 0.0) for r in exp_rows }

    # ----- Net per month (USD - USD) -----
    net_usd = [ totals_usd[i] - float(exp_map.get(lbl, 0.0)) for i, lbl in enumerate(labels) ]

    # ----- Snapshot for selected month -----
    latest = {
        "month":       labels[sel_idx],
        "count":       invoice_counts[sel_idx],
        "total":       totals_usd[sel_idx],                     # USD
        "kwh":         kwh[sel_idx],
        "paid":        paid_usd[sel_idx],                       # USD
        "unpaid":      unpaid_usd[sel_idx],                     # USD
        "avg_invoice": avg_inv_usd[sel_idx],                    # USD
        "expenses":    float(exp_map.get(labels[sel_idx], 0.0)),# USD (no /rate)
        "net_total":   net_usd[sel_idx],                        # USD
    }

    return render_template('dashboards.html',
                           labels=labels,
                           invoice_counts=invoice_counts,
                           totals=totals_usd,    # USD
                           kwh=kwh,
                           paid=paid_usd,        # USD
                           unpaid=unpaid_usd,    # USD
                           avg_invoice=avg_inv_usd,  # USD
                           latest=latest,        # USD snapshot
                           start=start_str, end=end_str)



if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)



